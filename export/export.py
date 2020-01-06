# export to excel 2010
from openpyxl import Workbook
import csv
import datetime
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles.colors import RED
from openpyxl.styles.colors import GREEN
from openpyxl.styles.colors import BLACK
from openpyxl.styles.colors import WHITE
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
font_red 	= Font(color=RED)
font_green 	= Font(color=GREEN)

from openpyxl.styles import NamedStyle, Font, Border, Side
highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True, size=14, color=WHITE)
bd = Side(style='thin', color="000000")
# bd = Side(color="000000")
highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
# highlight.fill = PatternFill("solid", fgColor="DDDDDD")
# highlight.fill = PatternFill("solid", fgColor="339966")
highlight.fill = PatternFill("solid", fgColor="4472c4")


# fill = PatternFill(fill_type=None,
#                 start_color='FFFFFFFF',
#                 end_color='FF000000')

wb = Workbook()
wb.add_named_style(highlight)

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
# ws['A1'] = 42

# Rows can also be appended
datetime_str = str(datetime.datetime.now())
ws.append(['Report generated Time', datetime_str])

# Python types will automatically be converted
# ws['A2'] = datetime.datetime.now()
ws.append([])
ws.append(['Table', 'Testcase', 'Result', 'Issue' ,'Details'])
ws['A3'].style = highlight
ws['B3'].style = highlight
ws['C3'].style = highlight
ws['D3'].style = highlight
ws['E3'].style = highlight

table = ''
testcase = ''
result = ''
details = ''
issue = ''

table_old = ''
testcase_old = ''
result_old = ''

textfile = "export/report-log.txt"
# textfile = "export/abcxyz.txt"

end_table_text = '============================================='

f = open(textfile, 'r')
reader = csv.reader(f, delimiter='\t')
for row in reader:
	if '======= test_result_of: ' in str(row):
		table_raw = str(row).replace('======= test_result_of: ','')
		table_raw = table_raw.replace('=','')
		table = table_raw.replace(' ','')
		table = table[2:-2]
	elif '--TC_' in str(row):
		testcase = str(row).replace('-','')
		testcase = testcase[2:-2]
	elif ('PASS' in str(row)) or ('FAIL' in str(row)):
		result = str(row)[2:-2]
	elif ( end_table_text not in str(row)) and row != []:
		details = details + str(row)[2:-2] + '\n'
		if '[Microsoft]' in str(row):
			issue = issue + str(row).split('[SQL Server]')[1].split('. ')[0] + '\n'
	if(testcase != testcase_old) or ( end_table_text in str(row)):
		if '--TC_001' not in str(row):
			# print('table = ' + table[2:-2])
			# print('testcase = ' + testcase_old)
			# print('result = ' + result)
			# print('details = ' + details)
			#update result
			if 'TC_001' in testcase_old:
				issue = details
				details = ''
			ws.append([table,testcase_old,result,issue,details])
		testcase_old = testcase
		result = ''
		details = ''
		issue = ''
	if(table != table_old):
		ws.append([])
		table_old = table

for i in range(5, 300):
	sheet = 'C' + str(i)
	result_ws = ws[sheet]
	if result_ws.value == 'PASS':
		result_ws.font = font_green
	elif result_ws.value == 'FAIL':
		result_ws.font = font_red

# Save the file
wb.save("export/report-" + datetime_str + ".xlsx")