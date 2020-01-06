from pandas import *
import openpyxl
from itertools import izip
fname = 'HM_DW_HIS_MappingFile_25122019.xlsx'
wb = openpyxl.load_workbook(fname)


sheet_names=['DIM_ICD'
,'DIM_DonViHanhChinh'
,'DIM_TienTe'
,'DIM_QuocGia'
,'DIM_Duoc_DonGia'
,'DIM_CheDoAnUong'#khac
,'DIM_CheDoChamSoc'
,'DIM_LyDoTiepNhan'
,'DIM_HinhThucDenKham'
,'DIM_NgheNghiep'
,'DIM_DanToc'
,'DIM_PhanLoaiKetQua'
,'DIM_GiaiQuyetKhamBenh'#khac
,'DIM_LyDoNhapVien'#khac
,'DIM_LoaiBenhAn'#khac
,'DIM_KipMo_VaiTro'#khac
,'DIM_TienTe_TyGia'
,'DIM_DuocCoSoThuoc'
]


for sheet_name in sheet_names:
	sheet = wb.get_sheet_by_name(sheet_name)
	# print('Please check again .... Often -1 for Source_Key.  Col =', sheet.max_row)
	num_start=3	
	num_end=sheet.max_row - 1

	# Target field
	x=[]
	for rowOfCellObjects in sheet['C%s'%num_start:'C%s'%num_end] :
	  	for cellObj in rowOfCellObjects:
	  		a=cellObj.value
	  		if a==u'\xa0':
	  			a = a.replace(u'\xa0', u'None')
	  		x.append(a)

	# Source Talbe
	n=[]
	for rowOfCellObjects0 in sheet['E%s'%num_start:'E%s'%num_end]:
		for cellObj0 in rowOfCellObjects0:
			d=cellObj0.value
			if d==u'\xa0':
				d = d.replace(u'\xa0', u'None')
			n.append(d)

	#Source field
	y=[]
	for rowOfCellObjects1 in sheet['F%s'%num_start:'F%s'%num_end]:
		for cellObj1 in rowOfCellObjects1:
			b=cellObj1.value
			if b==u'\xa0':
				b = b.replace(u'\xa0', u'None')
			y.append(b)

	#Reference table
	z=[]
	for rowOfCellObjects2 in sheet['G%s'%num_start:'G%s'%num_end]:
		for cellObj2 in rowOfCellObjects2:
			c=cellObj2.value
			if c==u'\xa0':
				c = c.replace(u'\xa0', u'None')
			z.append(c)

	#Foreign Key
	q=[]
	for rowOfCellObjects3 in sheet['H%s'%num_start:'H%s'%num_end]:
		for cellObj3 in rowOfCellObjects3:
			k=cellObj3.value
			if k==u'\xa0':
				k = k.replace(u'\xa0', u'None')
			q.append(k)


	d=0
	r=[]
	name="%s=["%(sheet_name)
	print(name)
	for l in range(len(x)):
		m=("('%s',               '%s',               '%s',                '%s',                    '%s',                 %s,                   %s)," %(x[d],n[d],y[d],z[d],q[d],None,None))
		m=m.replace("'None'","None")
		print(m)
		d=d+1
	print("('Hospital_Key',None,None,None,None,None,None,)" )
	print(']')
	print('\n')